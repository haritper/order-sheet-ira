from __future__ import annotations

import re
from dataclasses import dataclass
from functools import lru_cache
from pathlib import Path
from typing import Any

from flask import current_app
from openpyxl import load_workbook

from .db import execute, execute_many
from .fx import convert_source_to_display, get_daily_fx_snapshot


COST_HEADERS = {
    "fabric cost",
    "fabric cost avg",
    "printing cost",
    "printing cost avg",
    "c",
    "m",
    "t",
}

PRODUCT_SHEETS = {
    "MENS-T-SHIRT",
    "JACKETS",
    "TROUSERS",
    "CAPS",
    "CLADS",
    "T-SHIRT(WOMENS)",
    "W-JACKETS",
    "W-TROUSERS",
    "T-SHIRT(YOUTH)",
    "Y-JACKETS",
    "Y-TROUSERS",
}

SIZE_SEQUENCE = [
    "XS",
    "S",
    "M",
    "L",
    "XL",
    "2XL",
    "3XL",
    "4XL",
    "WXS",
    "WS",
    "WM",
    "WL",
    "WXL",
    "W2XL",
    "W3XL",
    "W4XL",
    "YXXS",
    "YXS",
    "YS",
    "YM",
    "YL",
    "YXL",
]


def slugify(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", value.lower()).strip()


def parse_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    return None


def parse_labeled_number(value: Any) -> tuple[str, float] | None:
    if not isinstance(value, str):
        return None
    match = re.search(r"^\s*([A-Za-z ]+?)\s*=?\s*([0-9]+(?:\.[0-9]+)?)\s*$", value.strip())
    if not match:
        return None
    return match.group(1).strip(), float(match.group(2))


@dataclass
class SheetView:
    worksheet: Any

    def __post_init__(self) -> None:
        self._merged_lookup: dict[str, str] = {}
        for merged_range in self.worksheet.merged_cells.ranges:
            start = merged_range.start_cell.coordinate
            for row in self.worksheet[merged_range.coord]:
                for cell in row:
                    self._merged_lookup[cell.coordinate] = start

    def actual(self, row: int, column: int) -> Any:
        return self.worksheet.cell(row=row, column=column).value

    def visible(self, row: int, column: int) -> Any:
        cell = self.worksheet.cell(row=row, column=column)
        source = self._merged_lookup.get(cell.coordinate)
        if source:
            return self.worksheet[source].value
        return cell.value


def find_header_row(sheet: SheetView) -> int | None:
    for row in range(1, min(sheet.worksheet.max_row, 5) + 1):
        for col in range(1, min(sheet.worksheet.max_column, 20) + 1):
            value = sheet.visible(row, col)
            if isinstance(value, str) and "PRODUCT CODE" in value.upper():
                return row
    return None


def header_map(sheet: SheetView, header_row: int) -> dict[str, int]:
    result: dict[str, int] = {}
    for col in range(1, sheet.worksheet.max_column + 1):
        value = sheet.visible(header_row, col)
        if isinstance(value, str) and value.strip():
            result[slugify(value)] = col
    return result


def first_visible_value(sheet: SheetView, column: int | None, start: int, end: int) -> Any:
    if not column:
        return None
    for row in range(start, end + 1):
        value = sheet.visible(row, column)
        if value not in (None, ""):
            return value
    return None


def category_for_sheet(sheet_name: str) -> str:
    name = sheet_name.upper()
    if "T-SHIRT" in name:
        return "shirt"
    if "JACKET" in name:
        return "jacket"
    if "TROUSER" in name:
        return "trouser"
    if "CAP" in name:
        return "cap"
    if "CLAD" in name:
        return "clad"
    return "other"


def infer_category_from_master(product_code: str, descriptor: Any) -> str:
    code = str(product_code or "").strip().upper()
    descriptor_text = slugify(str(descriptor or ""))
    prefix = code.split("-")[0] if code else ""
    if prefix in {"CT"}:
        return "cap"
    if "CLAD" in descriptor_text:
        return "clad"
    if "PANT" in descriptor_text or prefix in {"TP"}:
        return "trouser"
    if prefix in {"TJ", "UJ", "TOJ", "VJ", "WL", "PH", "SS", "JG", "ZH"}:
        return "jacket"
    return "shirt"


def infer_variant_from_master(product_code: str) -> str | None:
    code = str(product_code or "").upper()
    if "-HS-" in code:
        return "HALF SLEEVE"
    if "-FS-" in code:
        return "FULL SLEEVE"
    return None


def split_lines(value: Any) -> list[str]:
    if not isinstance(value, str):
        return []
    return [line.strip() for line in value.splitlines() if line.strip()]


def collect_accessories(
    sheet: SheetView,
    qty_col: int | None,
    rate_col: int | None,
    start: int,
    end: int,
) -> tuple[list[dict[str, Any]], float]:
    rows: list[dict[str, Any]] = []
    total = 0.0
    if not qty_col or not rate_col:
        return rows, total
    for row in range(start, end + 1):
        qty = parse_float(sheet.actual(row, qty_col))
        rate = parse_float(sheet.actual(row, rate_col))
        if qty is None and rate is None:
            continue
        line_total = round((qty or 0) * (rate or 0), 2)
        rows.append(
            {
                "row": row,
                "qty": qty or 0,
                "rate": rate or 0,
                "line_total": line_total,
            }
        )
        total += line_total
    return rows, round(total, 2)


def collect_costs(
    sheet: SheetView,
    columns: dict[str, int],
    start: int,
    end: int,
) -> tuple[dict[str, float], list[dict[str, Any]], float]:
    component_totals = {
        "fabric_cost": 0.0,
        "printing_cost": 0.0,
        "component_c": 0.0,
        "component_t": 0.0,
        "machine_cost": 0.0,
    }
    labor_breakdown: list[dict[str, Any]] = []
    total = 0.0

    for header_key, column in columns.items():
        if header_key not in COST_HEADERS:
            continue
        for row in range(start, end + 1):
            raw = sheet.actual(row, column)
            numeric = parse_float(raw)
            if numeric is not None:
                if header_key == "fabric cost" or header_key == "fabric cost avg":
                    component_totals["fabric_cost"] += numeric
                elif header_key == "printing cost" or header_key == "printing cost avg":
                    component_totals["printing_cost"] += numeric
                elif header_key == "c":
                    component_totals["component_c"] += numeric
                elif header_key == "t":
                    component_totals["component_t"] += numeric
                elif header_key == "m":
                    component_totals["machine_cost"] += numeric
                    labor_breakdown.append(
                        {
                            "row": row,
                            "header": header_key,
                            "label": "M",
                            "value": numeric,
                        }
                    )
                total += numeric
                continue

            parsed = parse_labeled_number(raw)
            if parsed is None:
                continue
            label, value = parsed
            labor_breakdown.append(
                {
                    "row": row,
                    "header": header_key,
                    "label": label,
                    "value": value,
                }
            )
            if header_key == "m":
                component_totals["machine_cost"] += value
            elif header_key == "c":
                component_totals["component_c"] += value
            elif header_key == "t":
                component_totals["component_t"] += value
            total += value

    rounded = {key: round(value, 2) for key, value in component_totals.items()}
    return rounded, labor_breakdown, round(total, 2)


def parse_product_sheet(sheet_name: str, worksheet) -> list[dict[str, Any]]:
    sheet = SheetView(worksheet)
    header_row = find_header_row(sheet)
    if not header_row:
        return []
    headers = header_map(sheet, header_row)

    code_col = headers.get("product code")
    if not code_col:
        return []

    anchors: list[int] = []
    for row in range(header_row + 1, worksheet.max_row + 1):
        raw_value = sheet.actual(row, code_col)
        if not isinstance(raw_value, str):
            continue
        text = raw_value.strip()
        if not text:
            continue
        upper = text.upper()
        if "PRODUCT CODE" in upper:
            continue
        anchors.append(row)

    rules: list[dict[str, Any]] = []
    qty_col = headers.get("qty")
    rate_col = headers.get("rate")
    accessory_col = headers.get("accessories")
    descriptor_col = headers.get("collar") or headers.get("styles") or headers.get("collar type")
    variant_col = headers.get("sleeeve type")
    unique_id_col = headers.get("unique id")
    product_col = headers.get("product")
    fabric_code_col = headers.get("fabric code")

    for index, start_row in enumerate(anchors):
        end_row = (anchors[index + 1] - 1) if index + 1 < len(anchors) else worksheet.max_row
        product_code = str(sheet.actual(start_row, code_col)).strip()
        accessory_breakdown, accessory_total = collect_accessories(
            sheet, qty_col, rate_col, start_row, end_row
        )
        component_totals, labor_breakdown, cost_total = collect_costs(
            sheet, headers, start_row, end_row
        )
        calculated_unit_rate = round(cost_total + accessory_total, 2)

        rules.append(
            {
                "rule_key": build_rule_key(
                    sheet_name,
                    product_code,
                    first_visible_value(sheet, descriptor_col, start_row, end_row),
                    first_visible_value(sheet, variant_col, start_row, end_row),
                    start_row,
                ),
                "product_code": product_code,
                "sheet_name": sheet_name,
                "category": category_for_sheet(sheet_name),
                "product_name": first_visible_value(sheet, product_col, start_row, end_row),
                "descriptor": first_visible_value(sheet, descriptor_col, start_row, end_row),
                "variant": first_visible_value(sheet, variant_col, start_row, end_row),
                "unique_id": first_visible_value(sheet, unique_id_col, start_row, end_row),
                "fabric_code": first_visible_value(sheet, fabric_code_col, start_row, end_row),
                "fabric_cost": component_totals["fabric_cost"],
                "printing_cost": component_totals["printing_cost"],
                "component_c": component_totals["component_c"],
                "component_t": component_totals["component_t"],
                "machine_cost": component_totals["machine_cost"],
                "accessory_total": accessory_total,
                "calculated_unit_rate": calculated_unit_rate,
                "accessory_breakdown": {
                    "names": split_lines(first_visible_value(sheet, accessory_col, start_row, end_row)),
                    "rows": accessory_breakdown,
                },
                "labor_breakdown": labor_breakdown,
                "source_row_start": start_row,
                "source_row_end": end_row,
            }
        )

    return rules


def parse_datasheet(workbook_path: Path) -> list[dict[str, Any]]:
    master_rules = parse_master_sheet(workbook_path)
    if master_rules:
        return master_rules

    workbook = load_workbook(workbook_path, data_only=False)
    rules: list[dict[str, Any]] = []
    for sheet_name in workbook.sheetnames:
        if sheet_name not in PRODUCT_SHEETS:
            continue
        rules.extend(parse_product_sheet(sheet_name, workbook[sheet_name]))
    return rules


def parse_master_sheet(workbook_path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(workbook_path, data_only=True)
    master_sheet = None
    for candidate in ("MASTER SHEET", "MASETR SHEET", "MASETR SHEET "):
        if candidate in workbook.sheetnames:
            master_sheet = workbook[candidate]
            break
    if master_sheet is None:
        return []

    grouped: dict[tuple[str, str, str, str, str], dict[str, Any]] = {}
    for row in range(2, master_sheet.max_row + 1):
        product_code = value_or_none(master_sheet.cell(row=row, column=1).value)
        if not isinstance(product_code, str):
            continue
        descriptor = value_or_none(master_sheet.cell(row=row, column=2).value)
        fabric_code = value_or_none(master_sheet.cell(row=row, column=3).value)
        size_raw = master_sheet.cell(row=row, column=4).value
        size = str(size_raw).strip().upper() if isinstance(size_raw, str) else None
        if size and size in {"XXL", "XXXL", "XXXXL"}:
            size = {"XXL": "2XL", "XXXL": "3XL", "XXXXL": "4XL"}[size]

        fabric_cost = parse_float(master_sheet.cell(row=row, column=5).value) or 0.0
        printing_cost = parse_float(master_sheet.cell(row=row, column=6).value) or 0.0
        cmt_cost = parse_float(master_sheet.cell(row=row, column=7).value) or 0.0
        accessories_cost = parse_float(master_sheet.cell(row=row, column=8).value) or 0.0
        design_cost = parse_float(master_sheet.cell(row=row, column=9).value) or 0.0
        order_sheet_cost = parse_float(master_sheet.cell(row=row, column=10).value) or 0.0
        total_cost = parse_float(master_sheet.cell(row=row, column=11).value)
        if total_cost is None:
            total_cost = round(
                fabric_cost
                + printing_cost
                + cmt_cost
                + accessories_cost
                + design_cost
                + order_sheet_cost,
                2,
            )
        else:
            total_cost = round(float(total_cost), 2)

        category = infer_category_from_master(product_code, descriptor)
        variant = infer_variant_from_master(product_code)
        key = (
            slugify(product_code),
            slugify(str(descriptor or "")),
            slugify(str(variant or "")),
            category,
            slugify(str(fabric_code or "")),
        )
        entry = grouped.get(key)
        if entry is None:
            entry = {
                "sheet_name": master_sheet.title.strip() or "MASTER SHEET",
                "product_code": product_code.strip(),
                "category": category,
                "product_name": descriptor,
                "descriptor": descriptor,
                "variant": variant,
                "unique_id": None,
                "fabric_code": fabric_code,
                "rows": [],
                "source_row_start": row,
                "source_row_end": row,
            }
            grouped[key] = entry
        entry["rows"].append(
            {
                "size": size,
                "fabric_cost": float(fabric_cost),
                "printing_cost": float(printing_cost),
                "cmt_cost": float(cmt_cost),
                "accessories_cost": float(accessories_cost),
                "design_cost": float(design_cost),
                "order_sheet_cost": float(order_sheet_cost),
                "total_cost": float(total_cost),
            }
        )
        entry["source_row_start"] = min(entry["source_row_start"], row)
        entry["source_row_end"] = max(entry["source_row_end"], row)

    rules: list[dict[str, Any]] = []
    for entry in grouped.values():
        rows = entry["rows"]
        count = max(len(rows), 1)
        avg_fabric = round(sum(row["fabric_cost"] for row in rows) / count, 2)
        avg_print = round(sum(row["printing_cost"] for row in rows) / count, 2)
        avg_cmt = round(sum(row["cmt_cost"] for row in rows) / count, 2)
        avg_accessories = round(sum(row["accessories_cost"] for row in rows) / count, 2)
        avg_design_and_order = round(
            (sum(row["design_cost"] for row in rows) + sum(row["order_sheet_cost"] for row in rows)) / count,
            2,
        )
        avg_total = round(sum(row["total_cost"] for row in rows) / count, 2)
        rules.append(
            {
                "rule_key": build_rule_key(
                    entry["sheet_name"],
                    entry["product_code"],
                    entry["descriptor"],
                    entry["variant"],
                    entry["source_row_start"],
                ),
                "product_code": entry["product_code"],
                "sheet_name": entry["sheet_name"],
                "category": entry["category"],
                "product_name": entry["product_name"],
                "descriptor": entry["descriptor"],
                "variant": entry["variant"],
                "unique_id": entry["unique_id"],
                "fabric_code": entry["fabric_code"],
                "fabric_cost": avg_fabric,
                "printing_cost": avg_print,
                "component_c": 0.0,
                "component_t": avg_design_and_order,
                "machine_cost": avg_cmt,
                "accessory_total": avg_accessories,
                "calculated_unit_rate": avg_total,
                "accessory_breakdown": {"names": [], "rows": []},
                "labor_breakdown": [],
                "source_row_start": entry["source_row_start"],
                "source_row_end": entry["source_row_end"],
            }
        )
    return rules


def import_pricing_rules(workbook_path: Path) -> int:
    rules = parse_datasheet(workbook_path)
    if not rules:
        return 0

    execute("DELETE FROM pricing_rules")
    execute_many(
        """
        INSERT INTO pricing_rules (
            rule_key, product_code, sheet_name, category, product_name, descriptor, variant,
            unique_id, fabric_code, fabric_cost, printing_cost, component_c,
            component_t, machine_cost, accessory_total, calculated_unit_rate,
            accessory_breakdown_json, labor_breakdown_json, source_row_start,
            source_row_end
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        [
            (
                rule["rule_key"],
                rule["product_code"],
                rule["sheet_name"],
                rule["category"],
                value_or_none(rule["product_name"]),
                value_or_none(rule["descriptor"]),
                value_or_none(rule["variant"]),
                value_or_none(rule["unique_id"]),
                value_or_none(rule["fabric_code"]),
                rule["fabric_cost"],
                rule["printing_cost"],
                rule["component_c"],
                rule["component_t"],
                rule["machine_cost"],
                rule["accessory_total"],
                rule["calculated_unit_rate"],
                json_dump(rule["accessory_breakdown"]),
                json_dump(rule["labor_breakdown"]),
                rule["source_row_start"],
                rule["source_row_end"],
            )
            for rule in rules
        ],
    )
    return len(rules)


def json_dump(value: Any) -> str:
    import json

    return json.dumps(value, ensure_ascii=True)


def value_or_none(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        return value.strip() or None
    return value


def list_pricing_rules() -> list[dict[str, Any]]:
    rows = execute(
        """
        SELECT
            *
        FROM pricing_rules
        ORDER BY category, sheet_name, product_code
        """
    ).fetchall()
    snapshot = get_daily_fx_snapshot()
    parsed_rows: list[dict[str, Any]] = []
    for row in rows:
        item = dict(row)
        item["calculated_unit_rate_inr"] = float(item["calculated_unit_rate"] or 0)
        item["effective_unit_rate_inr"] = (
            float(item["override_unit_rate"])
            if item["override_unit_rate"] is not None
            else item["calculated_unit_rate_inr"]
        )
        item["calculated_unit_rate_usd"] = convert_source_to_display(
            item["calculated_unit_rate_inr"], snapshot
        )
        item["effective_unit_rate_usd"] = convert_source_to_display(
            item["effective_unit_rate_inr"], snapshot
        )
        parsed_rows.append(item)
    return parsed_rows


@lru_cache(maxsize=4)
def workbook_size_references(workbook_path: str, modified_time: float) -> dict[str, Any]:
    workbook = load_workbook(workbook_path, data_only=True)

    def find_sheet(*candidates: str):
        for name in candidates:
            if name in workbook.sheetnames:
                return workbook[name]
        normalized_candidates = [slugify(name).replace(" ", "") for name in candidates]
        for sheet_name in workbook.sheetnames:
            normalized_sheet = slugify(sheet_name).replace(" ", "")
            if any(
                candidate in normalized_sheet or normalized_sheet in candidate
                for candidate in normalized_candidates
            ):
                return workbook[sheet_name]
        return None

    def normalize_size(value: Any) -> str | None:
        if not isinstance(value, str):
            return None
        text = value.strip().upper().replace(" ", "")
        aliases = {"XXL": "2XL", "XXXL": "3XL", "XXXXL": "4XL"}
        text = aliases.get(text, text)
        return text if text in SIZE_SEQUENCE else None

    fabric_rates: dict[str, float] = {}
    fabric_sheet = find_sheet("FABRIC")
    if fabric_sheet is not None:
        for row in range(2, fabric_sheet.max_row + 1):
            code = fabric_sheet[f"A{row}"].value
            price_per_kg = fabric_sheet[f"F{row}"].value
            if isinstance(code, str) and isinstance(price_per_kg, (int, float)):
                fabric_rates[code.strip()] = float(price_per_kg)

    top_fabric_costs: dict[str, dict[str, float]] = {}
    fabric_consumption_sheet = find_sheet(
        "FABRIC CONSUMPTION",
        "FABRIC CONSUMPTION (MENS)",
        "FABRIC WEIGHTS",
    )
    if fabric_consumption_sheet is not None:
        # Legacy format: size in column B, INS costs in E/F.
        for row in range(2, fabric_consumption_sheet.max_row + 1):
            size = normalize_size(fabric_consumption_sheet[f"B{row}"].value)
            if not size:
                continue
            half_cost = fabric_consumption_sheet[f"E{row}"].value
            full_cost = fabric_consumption_sheet[f"F{row}"].value
            if isinstance(half_cost, (int, float)) or isinstance(full_cost, (int, float)):
                top_fabric_costs[size] = {
                    "HALF SLEEVE": float(half_cost or 0),
                    "FULL SLEEVE": float(full_cost or 0),
                }

        # New mens format: size in A (or I due merged layouts), INS costs in D/E.
        if not top_fabric_costs:
            for row in range(2, fabric_consumption_sheet.max_row + 1):
                size = normalize_size(fabric_consumption_sheet[f"A{row}"].value) or normalize_size(
                    fabric_consumption_sheet[f"I{row}"].value
                )
                if not size:
                    continue
                half_cost = fabric_consumption_sheet[f"D{row}"].value
                full_cost = fabric_consumption_sheet[f"E{row}"].value
                if isinstance(half_cost, (int, float)) or isinstance(full_cost, (int, float)):
                    top_fabric_costs[size] = {
                        "HALF SLEEVE": float(half_cost or 0),
                        "FULL SLEEVE": float(full_cost or 0),
                    }

    mens_printing: dict[str, float] = {}
    printing_sheet = find_sheet("PRINTING")
    if printing_sheet is not None:
        for row in range(4, printing_sheet.max_row + 1):
            size = normalize_size(printing_sheet[f"E{row}"].value)
            ink_rate = printing_sheet[f"H{row}"].value
            if size and isinstance(ink_rate, (int, float)):
                mens_printing[size] = float(ink_rate)

    master_size_costs: dict[str, dict[str, dict[str, float]]] = {}
    master_sheet = find_sheet("MASTER SHEET", "MASETR SHEET", "MASETR SHEET ")
    if master_sheet is not None:
        for row in range(2, master_sheet.max_row + 1):
            raw_code = master_sheet[f"A{row}"].value
            if not isinstance(raw_code, str) or not raw_code.strip():
                continue
            product_code = raw_code.strip()
            size = normalize_size(master_sheet[f"D{row}"].value)
            if not size:
                continue
            fabric_cost = parse_float(master_sheet[f"E{row}"].value) or 0.0
            printing_cost = parse_float(master_sheet[f"F{row}"].value) or 0.0
            cmt_cost = parse_float(master_sheet[f"G{row}"].value) or 0.0
            accessories_cost = parse_float(master_sheet[f"H{row}"].value) or 0.0
            design_cost = parse_float(master_sheet[f"I{row}"].value) or 0.0
            order_sheet_cost = parse_float(master_sheet[f"J{row}"].value) or 0.0
            total_cost = parse_float(master_sheet[f"K{row}"].value)
            if total_cost is None:
                total_cost = (
                    fabric_cost
                    + printing_cost
                    + cmt_cost
                    + accessories_cost
                    + design_cost
                    + order_sheet_cost
                )
            master_size_costs.setdefault(product_code, {})[size] = {
                "unit_cost_inr": round(float(total_cost), 2),
                "fabric_cost_inr": round(float(fabric_cost), 2),
                "printing_cost_inr": round(float(printing_cost), 2),
                "fixed_cost_inr": round(float(cmt_cost + accessories_cost + design_cost + order_sheet_cost), 2),
            }

    return {
        "fabric_rates": fabric_rates,
        "top_fabric_costs": top_fabric_costs,
        "mens_printing": mens_printing,
        "master_size_costs": master_size_costs,
    }


def get_workbook_size_references() -> dict[str, Any]:
    workbook_path = Path(current_app.config["PRICING_WORKBOOK_PATH"]).resolve()
    modified_time = workbook_path.stat().st_mtime
    return workbook_size_references(str(workbook_path), modified_time)


def nearest_supported_size(size: str, available: dict[str, Any]) -> str | None:
    if size in available:
        return size
    if not available:
        return None
    if size in SIZE_SEQUENCE:
        target_index = SIZE_SEQUENCE.index(size)
        for index in range(target_index, -1, -1):
            candidate = SIZE_SEQUENCE[index]
            if candidate in available:
                return candidate
        for index in range(target_index + 1, len(SIZE_SEQUENCE)):
            candidate = SIZE_SEQUENCE[index]
            if candidate in available:
                return candidate
    return next(iter(available.keys()))


def size_factor_from_references(
    size: str,
    sleeve: str,
    base_fabric_rates: dict[str, dict[str, float]],
    printing_rates: dict[str, float],
) -> float:
    factors: list[float] = []
    if printing_rates:
        positive_printing = [value for value in printing_rates.values() if value > 0]
        if positive_printing:
            avg_print = sum(positive_printing) / len(positive_printing)
            matched = nearest_supported_size(size, printing_rates)
            size_print = float(printing_rates.get(matched or "", avg_print))
            if avg_print > 0:
                factors.append(size_print / avg_print)
    if base_fabric_rates:
        available_fabric = {
            key: float(value.get(sleeve, 0))
            for key, value in base_fabric_rates.items()
            if float(value.get(sleeve, 0)) > 0
        }
        if available_fabric:
            avg_fabric = sum(available_fabric.values()) / len(available_fabric)
            matched = nearest_supported_size(size, available_fabric)
            size_fabric = float(available_fabric.get(matched or "", avg_fabric))
            if avg_fabric > 0:
                factors.append(size_fabric / avg_fabric)
    if not factors:
        return 1.0
    # Keep the variation reasonable for non-exact categories.
    blended = sum(factors) / len(factors)
    return min(max(blended, 0.85), 1.2)


def calculate_sizewise_item_cost(rule: dict[str, Any], sizes: dict[str, int]) -> dict[str, Any] | None:
    category = str(rule.get("category") or "").lower()
    if category in {"cap", "clad"}:
        return None
    if rule.get("override_unit_rate") is not None:
        return None
    total_qty = sum(int(value or 0) for value in sizes.values())
    if total_qty <= 0:
        return None
    references = get_workbook_size_references()
    master_size_costs = references.get("master_size_costs", {})
    base_fabric_rates = references["top_fabric_costs"]
    printing_rates = references["mens_printing"]
    if not printing_rates and not base_fabric_rates and not master_size_costs:
        return None

    sleeve = "FULL SLEEVE" if "FULL" in str(rule.get("variant") or "").upper() else "HALF SLEEVE"
    snapshot = get_daily_fx_snapshot()
    size_breakdown: dict[str, dict[str, float]] = {}
    total_inr = 0.0

    code = str(rule.get("product_code") or "").strip()
    master_by_size = master_size_costs.get(code, {})
    if master_by_size:
        for size, quantity in sizes.items():
            quantity_int = int(quantity or 0)
            if quantity_int <= 0:
                continue
            matched_size = nearest_supported_size(size, master_by_size)
            if not matched_size:
                continue
            detail = master_by_size[matched_size]
            unit_inr = round(float(detail.get("unit_cost_inr") or 0), 2)
            if unit_inr <= 0:
                continue
            line_inr = round(unit_inr * quantity_int, 2)
            total_inr += line_inr
            size_breakdown[size] = {
                "quantity": quantity_int,
                "matched_size": matched_size,
                "fabric_cost_inr": round(float(detail.get("fabric_cost_inr") or 0), 2),
                "printing_cost_inr": round(float(detail.get("printing_cost_inr") or 0), 2),
                "fixed_cost_inr": round(float(detail.get("fixed_cost_inr") or 0), 2),
                "unit_cost_inr": unit_inr,
                "line_cost_inr": line_inr,
                "unit_cost_usd": convert_source_to_display(unit_inr, snapshot),
                "line_cost_usd": convert_source_to_display(line_inr, snapshot),
            }
        if size_breakdown:
            total_inr = round(total_inr, 2)
            total_usd = convert_source_to_display(total_inr, snapshot)
            average_unit_inr = round(total_inr / total_qty, 2)
            average_unit_usd = convert_source_to_display(average_unit_inr, snapshot)
            return {
                "size_breakdown": size_breakdown,
                "total_inr": total_inr,
                "total_usd": total_usd,
                "average_unit_inr": average_unit_inr,
                "average_unit_usd": average_unit_usd,
                "fixed_components_inr": round(
                    sum(float(item.get("fixed_cost_inr") or 0) for item in size_breakdown.values())
                    / len(size_breakdown),
                    2,
                ),
                "fabric_ratio": 1.0,
            }

    if category == "shirt":
        ins_rate = references["fabric_rates"].get("INS")
        target_rate = references["fabric_rates"].get(rule.get("fabric_code") or "", ins_rate)
        ratio = (target_rate / ins_rate) if ins_rate and target_rate else 1.0
        fixed_components = (
            float(rule.get("component_c") or 0)
            + float(rule.get("machine_cost") or 0)
            + float(rule.get("component_t") or 0)
            + float(rule.get("accessory_total") or 0)
        )

        for size, quantity in sizes.items():
            if not quantity:
                continue
            supported_size = nearest_supported_size(size, base_fabric_rates)
            if not supported_size:
                continue
            fabric_base = base_fabric_rates[supported_size].get(sleeve, 0.0)
            printing_size = nearest_supported_size(size, printing_rates)
            printing_cost = float(printing_rates.get(printing_size or "", 0.0))
            fabric_cost = round(fabric_base * ratio, 2)
            unit_inr = round(fabric_cost + printing_cost + fixed_components, 2)
            line_inr = round(unit_inr * int(quantity), 2)
            total_inr += line_inr
            size_breakdown[size] = {
                "quantity": int(quantity),
                "matched_size": supported_size,
                "fabric_cost_inr": fabric_cost,
                "printing_cost_inr": round(printing_cost, 2),
                "fixed_cost_inr": round(fixed_components, 2),
                "unit_cost_inr": unit_inr,
                "line_cost_inr": line_inr,
                "unit_cost_usd": convert_source_to_display(unit_inr, snapshot),
                "line_cost_usd": convert_source_to_display(line_inr, snapshot),
            }
        if not size_breakdown:
            return None
        total_inr = round(total_inr, 2)
        total_usd = convert_source_to_display(total_inr, snapshot)
        average_unit_inr = round(total_inr / total_qty, 2)
        average_unit_usd = convert_source_to_display(average_unit_inr, snapshot)
        return {
            "size_breakdown": size_breakdown,
            "total_inr": total_inr,
            "total_usd": total_usd,
            "average_unit_inr": average_unit_inr,
            "average_unit_usd": average_unit_usd,
            "fixed_components_inr": round(fixed_components, 2),
            "fabric_ratio": round(ratio, 4),
        }

    # Trouser/jacket/etc. with size-wise distribution using workbook size factor.
    base_unit_inr = float(rule.get("effective_unit_rate_inr") or rule.get("calculated_unit_rate") or 0)
    if base_unit_inr <= 0:
        return None
    for size, quantity in sizes.items():
        quantity_int = int(quantity or 0)
        if quantity_int <= 0:
            continue
        factor = size_factor_from_references(size, sleeve, base_fabric_rates, printing_rates)
        unit_inr = round(base_unit_inr * factor, 2)
        line_inr = round(unit_inr * quantity_int, 2)
        total_inr += line_inr
        size_breakdown[size] = {
            "quantity": quantity_int,
            "matched_size": nearest_supported_size(size, printing_rates) or size,
            "fabric_cost_inr": 0.0,
            "printing_cost_inr": 0.0,
            "fixed_cost_inr": round(unit_inr, 2),
            "unit_cost_inr": unit_inr,
            "line_cost_inr": line_inr,
            "unit_cost_usd": convert_source_to_display(unit_inr, snapshot),
            "line_cost_usd": convert_source_to_display(line_inr, snapshot),
        }
    if not size_breakdown:
        return None

    total_inr = round(total_inr, 2)
    total_usd = convert_source_to_display(total_inr, snapshot)
    average_unit_inr = round(total_inr / total_qty, 2)
    average_unit_usd = convert_source_to_display(average_unit_inr, snapshot)
    return {
        "size_breakdown": size_breakdown,
        "total_inr": total_inr,
        "total_usd": total_usd,
        "average_unit_inr": average_unit_inr,
        "average_unit_usd": average_unit_usd,
        "fixed_components_inr": round(base_unit_inr, 2),
        "fabric_ratio": 1.0,
    }


def calculate_sizewise_shirt_cost(rule: dict[str, Any], sizes: dict[str, int]) -> dict[str, Any] | None:
    # Backward-compatible alias used by existing order builder code.
    return calculate_sizewise_item_cost(rule, sizes)


def update_pricing_override(rule_id: int, override_rate: float | None) -> None:
    execute(
        """
        UPDATE pricing_rules
        SET override_unit_rate = ?, updated_at = CURRENT_TIMESTAMP
        WHERE id = ?
        """,
        (override_rate, rule_id),
    )


def find_pricing_rule(category: str, descriptor: str | None = None, variant: str | None = None):
    rows = execute(
        """
        SELECT
            *
        FROM pricing_rules
        WHERE category = ?
        """,
        (category,),
    ).fetchall()
    if not rows:
        return None

    normalized_descriptor = slugify(descriptor or "")
    normalized_variant = slugify(variant or "")

    def score(row) -> tuple[int, int]:
        descriptor_text = slugify(row["descriptor"] or "")
        variant_text = slugify(row["variant"] or "")
        descriptor_score = 0
        variant_score = 0
        if normalized_descriptor:
            if normalized_descriptor in descriptor_text:
                descriptor_score = 3
            elif any(token and token in descriptor_text for token in normalized_descriptor.split()):
                descriptor_score = 2
            elif descriptor_alias_match(normalized_descriptor, descriptor_text):
                descriptor_score = 1
        if normalized_variant:
            if normalized_variant in variant_text:
                variant_score = 2
            elif any(token and token in variant_text for token in normalized_variant.split()):
                variant_score = 1
        return descriptor_score, variant_score

    sorted_rows = sorted(rows, key=score, reverse=True)
    selected = dict(sorted_rows[0])
    snapshot = get_daily_fx_snapshot()
    selected["calculated_unit_rate_inr"] = float(selected["calculated_unit_rate"] or 0)
    selected["effective_unit_rate_inr"] = (
        float(selected["override_unit_rate"])
        if selected["override_unit_rate"] is not None
        else selected["calculated_unit_rate_inr"]
    )
    selected["calculated_unit_rate_usd"] = convert_source_to_display(
        selected["calculated_unit_rate_inr"], snapshot
    )
    selected["effective_unit_rate_usd"] = convert_source_to_display(
        selected["effective_unit_rate_inr"], snapshot
    )
    return selected


def pricing_catalog_summary(pricing_rules: list[dict[str, Any]]) -> list[dict[str, Any]]:
    labels = {
        "shirt": "T Shirts",
        "trouser": "Trousers",
        "jacket": "Jackets",
        "cap": "Caps",
        "clad": "Clads",
    }
    ordered_categories = ["shirt", "trouser", "jacket", "cap", "clad"]
    grouped: list[dict[str, Any]] = []
    for category in ordered_categories:
        items = [rule for rule in pricing_rules if summary_rule_eligible(rule, category)]
        if not items:
            continue
        positive_inr = [
            float(rule["effective_unit_rate_inr"])
            for rule in items
            if float(rule["effective_unit_rate_inr"]) > 0
        ]
        positive_usd = [
            float(rule["effective_unit_rate_usd"])
            for rule in items
            if float(rule["effective_unit_rate_usd"]) > 0
        ]
        starts_from_inr = min(positive_inr) if positive_inr else 0.0
        starts_from_usd = min(positive_usd) if positive_usd else 0.0
        descriptor_count = len({(rule["descriptor"] or "").strip() for rule in items if rule["descriptor"]})
        variant_count = len({(rule["variant"] or "").strip() for rule in items if rule["variant"]})
        grouped.append(
            {
                "category": category,
                "label": labels.get(category, category.title()),
                "starts_from_inr": round(starts_from_inr, 2),
                "starts_from_usd": round(starts_from_usd, 2),
                "descriptor_count": descriptor_count,
                "variant_count": variant_count,
            }
        )
    return grouped


def summary_rule_eligible(rule: dict[str, Any], category: str) -> bool:
    if rule["category"] != category:
        return False
    effective_inr = float(rule.get("effective_unit_rate_inr") or 0)
    if effective_inr <= 0:
        return False
    sheet_name = str(rule.get("sheet_name") or "")
    upper_sheet = sheet_name.upper()
    if "MASTER" in upper_sheet or "MASETR" in upper_sheet:
        return effective_inr > 0
    if category == "shirt":
        return (
            sheet_name == "MENS-T-SHIRT"
            and float(rule.get("fabric_cost") or 0) > 0
            and float(rule.get("printing_cost") or 0) > 0
        )
    if category == "trouser":
        return sheet_name == "TROUSERS" and float(rule.get("fabric_cost") or 0) > 0
    if category == "jacket":
        return sheet_name == "JACKETS" and float(rule.get("fabric_cost") or 0) > 0
    if category == "cap":
        return sheet_name == "CAPS" and effective_inr > 0
    if category == "clad":
        return sheet_name == "CLADS" and effective_inr > 0
    return True


def descriptor_alias_match(needle: str, haystack: str) -> bool:
    aliases = {
        "single piping": ["piping pant", "pining pant"],
        "polo": ["polo"],
        "mandarin": ["mandarin", "mandrain"],
        "zip up": ["zip up", "zipup"],
        "insert": ["insert"],
        "rib": ["rib"],
        "top panel": ["top panel"],
        "bottom panel": ["bottom panel", "btm panel"],
        "top bottom": ["top bottom", "top btm"],
        "middle panel": ["middle panel"],
        "top stripe panel": ["top stripe panel"],
        "base": ["base"],
        "mixed": ["mixed"],
        "bottom": ["bottom"],
        "helmet clad": ["helmet clad"],
        "pad clad": ["pad clad"],
    }
    for key, values in aliases.items():
        if key in needle:
            return any(value in haystack for value in values)
    return False


def build_rule_key(
    sheet_name: str,
    product_code: str,
    descriptor: Any,
    variant: Any,
    start_row: int,
) -> str:
    parts = [
        slugify(sheet_name).replace(" ", "-"),
        slugify(product_code).replace(" ", "-"),
        slugify(str(descriptor or "na")).replace(" ", "-"),
        slugify(str(variant or "na")).replace(" ", "-"),
        str(start_row),
    ]
    return "|".join(parts)

