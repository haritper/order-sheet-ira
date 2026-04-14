import csv
import json
from pathlib import Path

from openpyxl import load_workbook


def load_input_rows(input_file, sheet_name=None):
    path = Path(input_file)
    ext = path.suffix.lower()

    if ext in {".csv"}:
        with path.open("r", encoding="utf-8", errors="replace") as f:
            rows = list(csv.DictReader(f))
        input_type = "csv"
        active_sheet = ""
    elif ext in {".json"}:
        payload = json.loads(path.read_text(encoding="utf-8"))
        if isinstance(payload, list):
            rows = payload
        else:
            rows = payload.get("rows") or payload.get("normalized_rows") or []
        input_type = "json"
        active_sheet = ""
    elif ext in {".xlsx", ".xlsm"}:
        wb = load_workbook(path, data_only=True)
        ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]
        active_sheet = ws.title
        values = list(ws.values)
        headers = [str(h).strip() if h is not None else "" for h in values[0]]
        rows = []
        for row in values[1:]:
            if not any(cell is not None and str(cell).strip() for cell in row):
                continue
            rows.append({headers[i]: row[i] for i in range(len(headers))})
        input_type = "excel"
    else:
        raise ValueError(f"Unsupported input extension: {ext}")

    return {
        "rows": rows,
        "input_type": input_type,
        "sheet_name": active_sheet,
        "input_file_name": path.name,
    }
