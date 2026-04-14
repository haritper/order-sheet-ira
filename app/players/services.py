import csv
import io
import json
import os
import re

from openpyxl import load_workbook

from app.models import SleeveType

ALLOWED_SIZES = {
    "XS",
    "S",
    "M",
    "L",
    "XL",
    "2XL",
    "3XL",
    "4XL",
    "WXS",
    "WS",
    "WM",
    "WL",
    "WXL",
    "W2XL",
    "W3XL",
    "W4XL",
    "YXXS",
    "YXS",
    "YS",
    "YM",
    "YL",
    "YXL",
}

REQUIRED_COLUMNS = [
    "player_name",
    "number",
    "sleeve_type",
    "tshirt_size",
    "tshirt_qty",
    "trouser_size",
    "trouser_qty",
]

HEADER_ALIASES = {
    "player_name": [
        "player_name",
        "name",
        "player",
        "full_name",
        "player name",
        "preferred_name_on_jersey",
        "preferred_jersey_name",
    ],
    "jersey_name": [
        "name_on_jersey",
        "display_name",
        "jersey_name",
        "name_to_be_printed",
        "preferred_name_on_jersey",
        "preferred_jersey_name",
    ],
    "first_name": ["player_first_name", "first_name", "player firstname", "firstname"],
    "last_name": ["player_last_name", "last_name", "player lastname", "lastname"],
    "category": ["category", "gender", "team_category", "group", "division"],
    "number": [
        "number",
        "number_on_jersey",
        "number on jersey",
        "no",
        "jersey_number",
        "player_number",
        "jersey no",
        "shirt no",
        "preferred_jersey_no",
        "preferred_jersey_number",
    ],
    "sleeve_type": [
        "sleeve_type",
        "sleeve",
        "sleeve type",
        "shirt_sleeve",
        "shirt sleeve",
        "jersey_sleeve",
        "jersey sleeve",
        "half_sleeve_full_sleeve",
        "long_or_short_sleeve",
    ],
    "tshirt_size": [
        "tshirt_size",
        "shirt_size",
        "t shirt size",
        "tee_size",
        "shirt size",
        "jersey_size_no_jersey",
        "jersey_size",
    ],
    "tshirt_qty": ["tshirt_qty", "shirt_qty", "t shirt qty", "tee_qty", "shirt qty", "quantity"],
    "trouser_size": [
        "trouser_size",
        "pant_size",
        "pants_size",
        "trouser size",
        "pant size",
        "pant_size_no_pants",
        "pants_pant_size_no_pants",
        "trouser_details_trouser_size",
        "travel_trouser_details_trouser_size",
    ],
    "trouser_qty": [
        "trouser_qty",
        "pant_qty",
        "pants_qty",
        "trouser qty",
        "pant qty",
        "qty",
        "pants_qty_1",
        "pants_qty_2",
        "pants_qty_3",
    ],
    "cap_qty": ["cap_qty", "cap_qty_1", "cap_qty_2", "cap qty"],
    "baggy_cap_qty": ["baggy_cap_qty", "baggy cap qty", "baggycap_qty", "baggy cap quantity"],
    "hat_qty": ["hat_qty", "hat qty", "hat quantity", "hats_qty", "cap_hat_hat_qty"],
    "pad_clad_qty": ["pad_clad_qty", "pad clad qty", "padclad_qty", "pad_clad quantity"],
    "helmet_clad_qty": [
        "helmet_clad_qty",
        "helmet clad qty",
        "helmetclad_qty",
        "helmet_clad quantity",
        "helmat_clad_qty",
    ],
    "row_number": ["row_number", "row", "s_no", "s.no", "serial no", "sl no"],
}

ALIAS_KEY_SET = {
    "".join(c.lower() if c.isalnum() else "_" for c in alias).strip("_")
    for aliases in HEADER_ALIASES.values()
    for alias in aliases
}


def parse_players_csv(file_storage):
    payload = file_storage.read().decode("utf-8", errors="replace")
    stream = io.StringIO(payload)
    reader = csv.DictReader(stream)

    missing = [col for col in REQUIRED_COLUMNS if col not in (reader.fieldnames or [])]
    if missing:
        return [], [{"row": 0, "error": f"Missing columns: {', '.join(missing)}"}]

    valid_rows = []
    errors = []

    for idx, row in enumerate(reader, start=1):
        row_data, row_errors = normalize_player_row(row, idx)
        if row_data is None and not row_errors:
            continue
        if row_errors:
            errors.append({"row": idx, "error": "; ".join(row_errors)})
        else:
            valid_rows.append(row_data)

    return valid_rows, errors


def parse_players_clean(file_storage):
    filename = (file_storage.filename or "").lower()
    if filename.endswith(".csv"):
        return parse_players_csv(file_storage)
    if filename.endswith(".xlsx"):
        blob = file_storage.read()
        raw_rows, extraction_error = _extract_tabular_rows(blob, filename)
        if extraction_error:
            return [], [{"row": 0, "error": extraction_error}]
        mapped_rows = [_map_raw_row(row) for row in raw_rows]
        mapped_rows = [r for r in mapped_rows if _is_roster_candidate(r)]
        return _validate_rows(mapped_rows)
    return [], [{"row": 0, "error": "Unsupported file type. Upload only .csv or .xlsx"}]


def build_error_csv(errors):
    stream = io.StringIO()
    writer = csv.DictWriter(stream, fieldnames=["row", "error"])
    writer.writeheader()
    for error in errors:
        writer.writerow(error)
    return stream.getvalue()


def normalize_player_row(row, idx):
    row_errors = []
    name = (row.get("player_name") or "").strip()
    number = _normalize_number(row.get("number"))
    sleeve = _normalize_sleeve(str(row.get("sleeve_type") or "").strip())
    t_size = _normalize_size(str(row.get("tshirt_size") or "").strip())
    tr_size = _normalize_size(str(row.get("trouser_size") or "").strip())

    # Read qtys early so zero-qty entries can be ignored instead of treated as errors.
    t_qty_raw = _parse_int_value(row.get("tshirt_qty"))
    tr_qty_raw = _parse_int_value(row.get("trouser_qty"))

    has_tshirt = bool(t_size)
    has_trouser = bool(tr_size)

    if has_tshirt and t_qty_raw is not None and t_qty_raw <= 0:
        has_tshirt = False
        t_size = ""
    if has_trouser and tr_qty_raw is not None and tr_qty_raw <= 0:
        has_trouser = False
        tr_size = ""

    # Skip row only when both tshirt and trouser sizes are missing.
    if not has_tshirt and not has_trouser:
        return None, []

    if not name:
        row_errors.append("player_name is required")
    if not number:
        number = "0"

    # Sleeve only matters for tshirt rows.
    if not sleeve and has_tshirt:
        # Default missing sleeve to HALF when size exists.
        sleeve = SleeveType.HALF.value
    if has_tshirt and sleeve not in {SleeveType.HALF.value, SleeveType.FULL.value, "3/4 TH"}:
        row_errors.append("sleeve_type must be HALF, FULL or 3/4 TH")
    if has_tshirt and t_size not in ALLOWED_SIZES:
        row_errors.append("tshirt_size is invalid")
    if has_trouser and tr_size not in ALLOWED_SIZES:
        row_errors.append("trouser_size is invalid")

    t_qty = t_qty_raw
    if has_tshirt:
        if t_qty is None:
            t_qty = 1
        if t_qty <= 0:
            row_errors.append("tshirt_qty must be > 0")
    else:
        t_qty = 0

    tr_qty = tr_qty_raw
    if has_trouser:
        if tr_qty is None:
            tr_qty = 1
        if tr_qty <= 0:
            row_errors.append("trouser_qty must be > 0")
    else:
        tr_qty = 0

    if not has_tshirt:
        sleeve = sleeve or SleeveType.HALF.value

    row_number_raw = str(row.get("row_number") or "").strip()
    try:
        row_number = int(row_number_raw) if row_number_raw else idx
    except ValueError:
        row_errors.append("row_number must be integer")
        row_number = idx

    row_data = {
        "row_number": row_number,
        "player_name": name,
        "number": number,
        "sleeve_type": sleeve,
        "tshirt_size": t_size,
        "tshirt_qty": t_qty,
        "trouser_size": tr_size,
        "trouser_qty": tr_qty,
    }
    return row_data, row_errors


def parse_players_ai(file_storage):
    filename = (file_storage.filename or "").lower()
    if not (filename.endswith(".xlsx") or filename.endswith(".csv")):
        raise ValueError("Unsupported file type for AI import. Upload only .xlsx or .csv")

    blob = file_storage.read()
    raw_rows, extraction_error = _extract_tabular_rows(blob, filename)
    if extraction_error:
        return [], [{"row": 0, "error": extraction_error}]

    mapped_rows = [_map_raw_row(row) for row in raw_rows]
    mapped_rows = [r for r in mapped_rows if _is_roster_candidate(r)]
    det_valid, det_errors = _validate_rows(mapped_rows)

    # Always trust deterministic parse when it produces at least one valid row.
    # This preserves explicit values like WM/WS/WL from sheet cells.
    if det_valid:
        return det_valid, det_errors

    # Otherwise attempt AI assist only when deterministic yielded no usable rows.
    extracted = {"kind": "text", "text": json.dumps(raw_rows), "source": filename}
    try:
        llm_rows = call_openai_for_roster(extracted)
    except RuntimeError:
        return det_valid, det_errors
    ai_valid, ai_errors = _validate_rows(llm_rows)

    if ai_valid:
        return ai_valid, ai_errors
    return det_valid, det_errors


def parse_accessory_totals(file_storage):
    filename = (file_storage.filename or "").lower()
    if not (filename.endswith(".xlsx") or filename.endswith(".csv")):
        return _empty_accessory_totals(), False

    blob = file_storage.read()
    raw_rows, extraction_error = _extract_tabular_rows(blob, filename)
    if extraction_error:
        return _empty_accessory_totals(), False

    mapped_rows = [_map_raw_row(row) for row in raw_rows]
    totals = _empty_accessory_totals()
    found_any_accessory_column = False

    for row in mapped_rows:
        cap_qty = _parse_int_value(row.get("cap_qty"))
        baggy_cap_qty = _parse_int_value(row.get("baggy_cap_qty"))
        hat_qty = _parse_int_value(row.get("hat_qty"))
        pad_clad_qty = _parse_int_value(row.get("pad_clad_qty"))
        helmet_clad_qty = _parse_int_value(row.get("helmet_clad_qty"))

        if cap_qty is not None:
            totals["CAP"] += max(0, cap_qty)
            found_any_accessory_column = True
        if baggy_cap_qty is not None:
            totals["BAGGY CAP"] += max(0, baggy_cap_qty)
            found_any_accessory_column = True
        if hat_qty is not None:
            totals["HAT"] += max(0, hat_qty)
            found_any_accessory_column = True
        if pad_clad_qty is not None:
            totals["PAD CLAD"] += max(0, pad_clad_qty)
            found_any_accessory_column = True
        if helmet_clad_qty is not None:
            totals["HELMET CLAD"] += max(0, helmet_clad_qty)
            found_any_accessory_column = True

    return totals, found_any_accessory_column


def parse_product_item_totals(file_storage):
    filename = (file_storage.filename or "").lower()
    if not (filename.endswith(".xlsx") or filename.endswith(".csv")):
        return {}, False

    blob = file_storage.read()
    raw_rows, extraction_error = _extract_tabular_rows(blob, filename)
    if extraction_error:
        return {}, False

    totals = {}
    found_any = False

    def _first_non_empty(raw, *keys):
        for key in keys:
            value = raw.get(key)
            if value is None:
                continue
            if str(value).strip() == "":
                continue
            return value
        return ""

    def _add(product_key, gender, sleeve, size_value, qty_value):
        nonlocal found_any
        qty = _parse_int_value(qty_value)
        size = _normalize_size(size_value or "")
        if qty is None or qty <= 0 or not size:
            return
        bucket = _item_size_bucket(size)
        if not bucket:
            return
        g = _normalize_gender(gender)
        s = _normalize_sleeve(sleeve or "")
        k = (product_key, g, s)
        row = totals.setdefault(
            k,
            {
                "qty_xs": 0,
                "qty_s": 0,
                "qty_m": 0,
                "qty_l": 0,
                "qty_xl": 0,
                "qty_2xl": 0,
                "qty_3xl": 0,
                "qty_4xl": 0,
            },
        )
        row[bucket] += qty
        found_any = True

    for raw in raw_rows:
        gender = _first_non_empty(raw, "gender", "category") or "MENS"
        playing_sleeve = _first_non_empty(
            raw,
            "playing_t_shirt_details_sleeve_type",
            "sleeve_type",
            "playing_t_shirt_sleeve_type",
            "t_shirt_sleeve_type",
        ) or "HALF"
        # Business rule: Training Jersey is HALF sleeve only.
        training_sleeve = "HALF"

        _add(
            "playing_jersey",
            gender,
            playing_sleeve,
            _first_non_empty(
                raw,
                "t_shirt_size",
                "playing_t_shirt_size",
                "playing_tshirt_size",
                "playing_jersey_tshirt_size",
            ),
            _first_non_empty(
                raw,
                "t_shirt_qty",
                "playing_t_shirt_qty",
                "playing_tshirt_qty",
                "playing_jersey_tshirt_qty",
            ),
        )
        _add(
            "training_jersey",
            gender,
            training_sleeve,
            _first_non_empty(
                raw,
                "training_t_shirt_details_t_shirt_size",
                "training_t_shirt_size",
                "training_tshirt_size",
            ),
            _first_non_empty(
                raw,
                "t_shirt_qty_1",
                "training_t_shirt_qty",
                "training_tshirt_qty",
            ),
        )
        _add(
            "trouser",
            gender,
            "",
            _first_non_empty(
                raw,
                "trouser_details_trouser_size",
                "trouser_size",
                "trousers_size",
            ),
            _first_non_empty(
                raw,
                "trouser_qty",
                "trousers_qty",
            ),
        )
        _add(
            "shorts",
            gender,
            "",
            _first_non_empty(
                raw,
                "shorts_details_shorts_size",
                "shorts_size",
            ),
            _first_non_empty(
                raw,
                "shorts_qty",
            ),
        )
        _add(
            "jacket",
            gender,
            "",
            _first_non_empty(
                raw,
                "travel_jacket_details_jacket_size",
                "jacket_size",
            ),
            _first_non_empty(
                raw,
                "jacket_qty",
            ),
        )
        _add(
            "travel_trouser",
            gender,
            "",
            _first_non_empty(
                raw,
                "travel_trouser_details_trouser_size",
                "travel_trousers_size",
                "travel_trouser_size",
            ),
            _first_non_empty(
                raw,
                "trouser_qty_1",
                "travel_trousers_qty",
                "travel_trouser_qty",
            ),
        )
        _add(
            "sleeveless_jacket",
            gender,
            "",
            _first_non_empty(
                raw,
                "sleeveless_jacket_details_jacket_size",
                "sleeveless_jacket_size",
            ),
            _first_non_empty(
                raw,
                "jacket_qty_1",
                "sleeveless_jacket_qty",
            ),
        )
        _add(
            "polo",
            gender,
            "",
            _first_non_empty(
                raw,
                "travel_polo_t_shirt_size",
                "travel_polo_size",
                "polo_size",
            ),
            _first_non_empty(
                raw,
                "t_shirt_qty_2",
                "travel_polo_qty",
                "polo_qty",
            ),
        )

    return totals, found_any


def _empty_accessory_totals():
    return {
        "CAP": 0,
        "BAGGY CAP": 0,
        "HAT": 0,
        "PAD CLAD": 0,
        "HELMET CLAD": 0,
    }


def _normalize_gender(value):
    txt = str(value or "").strip().upper()
    if "WOMEN" in txt:
        return "WOMENS"
    if "YOUTH" in txt or "KID" in txt:
        return "YOUTH"
    return "MENS"


def _item_size_bucket(size):
    s = str(size or "").strip().upper()
    womens = {"WXS": "XS", "WS": "S", "WM": "M", "WL": "L", "WXL": "XL", "W2XL": "2XL", "W3XL": "3XL", "W4XL": "4XL"}
    youth = {"YXXS": "XS", "YXS": "S", "YS": "M", "YM": "L", "YL": "XL", "YXL": "2XL"}
    s = womens.get(s, s)
    s = youth.get(s, s)
    mapping = {
        "XS": "qty_xs",
        "S": "qty_s",
        "M": "qty_m",
        "L": "qty_l",
        "XL": "qty_xl",
        "2XL": "qty_2xl",
        "3XL": "qty_3xl",
        "4XL": "qty_4xl",
    }
    return mapping.get(s)


def _validate_rows(rows):
    valid_rows = []
    errors = []
    for idx, row in enumerate(rows, start=1):
        normalized, row_errors = normalize_player_row(row, idx)
        if normalized is None and not row_errors:
            continue
        if row_errors:
            errors.append({"row": idx, "error": "; ".join(row_errors)})
        else:
            valid_rows.append(normalized)
    return valid_rows, errors


def _extract_tabular_rows(blob, filename):
    if filename.endswith(".csv"):
        payload = blob.decode("utf-8", errors="replace")
        matrix = list(csv.reader(io.StringIO(payload)))
        if not matrix:
            return [], "No rows found in CSV"
        rows, _, _ = _extract_rows_from_matrix(matrix, sheet_name="CSV")
        if not rows:
            return [], "Could not detect roster rows in CSV"
        return rows, None

    wb = load_workbook(io.BytesIO(blob), data_only=True)
    best = None

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        values = list(ws.values)
        if not values:
            continue

        candidate_rows, score, used_header_idx = _extract_rows_from_matrix(values, sheet_name)
        if not candidate_rows:
            continue
        mapped_rows = [_map_raw_row(row) for row in candidate_rows]
        mapped_rows = [row for row in mapped_rows if _is_roster_candidate(row)]
        det_valid, det_errors = _validate_rows(mapped_rows)
        duplicate_cols = _duplicate_column_count(candidate_rows)
        candidate = {
            "sheet": sheet_name,
            "header_idx": used_header_idx,
            "score": score,
            "rows": candidate_rows,
            "valid_count": len(det_valid),
            "error_count": len(det_errors),
            "duplicate_cols": duplicate_cols,
        }

        if best is None:
            best = candidate
        elif candidate["valid_count"] > best["valid_count"]:
            best = candidate
        elif candidate["valid_count"] == best["valid_count"] and candidate["error_count"] < best["error_count"]:
            best = candidate
        elif (
            candidate["valid_count"] == best["valid_count"]
            and candidate["error_count"] == best["error_count"]
            and candidate["header_idx"] is not None
            and best["header_idx"] is None
        ):
            best = candidate
        elif (
            candidate["valid_count"] == best["valid_count"]
            and candidate["error_count"] == best["error_count"]
            and candidate["header_idx"] is None
            and best["header_idx"] is not None
        ):
            # Prefer real header-detected tables over headerless fallbacks.
            pass
        elif (
            candidate["valid_count"] == best["valid_count"]
            and candidate["error_count"] == best["error_count"]
            and candidate["header_idx"] is not None
            and best["header_idx"] is not None
            and candidate["score"] > best["score"]
        ):
            best = candidate
        elif (
            candidate["valid_count"] == best["valid_count"]
            and candidate["error_count"] == best["error_count"]
            and candidate["duplicate_cols"] < best["duplicate_cols"]
        ):
            best = candidate
        elif (
            candidate["valid_count"] == best["valid_count"]
            and candidate["error_count"] == best["error_count"]
            and candidate["duplicate_cols"] == best["duplicate_cols"]
            and candidate["score"] > best["score"]
        ):
            best = candidate
        elif (
            candidate["valid_count"] == best["valid_count"]
            and candidate["error_count"] == best["error_count"]
            and candidate["duplicate_cols"] == best["duplicate_cols"]
            and candidate["score"] == best["score"]
            and len(candidate["rows"]) > len(best["rows"])
        ):
            best = candidate

    if not best or not best["rows"]:
        return [], "Could not detect a roster table in Excel (header row not found)"

    return best["rows"], None


def _duplicate_column_count(rows):
    if not rows:
        return 0
    sample = rows[0]
    return sum(1 for k in sample.keys() if isinstance(k, str) and k.endswith("_1"))


def _extract_rows_from_matrix(values, sheet_name):
    header_idx, score = _find_header_row(values)
    if header_idx is not None:
        group_idx = header_idx - 1 if header_idx > 0 else None
        if _should_shift_to_subheader(values, header_idx):
            group_idx = header_idx
            header_idx = header_idx + 1

        headers = _build_headers(values, header_idx, group_idx=group_idx)
        rows = []
        for excel_row_num, row in enumerate(values[header_idx + 1 :], start=header_idx + 2):
            if not any(cell is not None and str(cell).strip() for cell in row):
                continue
            record = {}
            all_values = []
            extra_context_values = []
            for i, header in enumerate(headers):
                cell_value = "" if i >= len(row) or row[i] is None else str(row[i]).strip()
                if cell_value:
                    all_values.append(cell_value)
                if not header:
                    if cell_value:
                        extra_context_values.append(cell_value)
                    continue
                record[header] = cell_value
            record["_source_excel_row"] = excel_row_num
            record["_source_sheet"] = sheet_name
            record["_raw_row_values"] = all_values
            record["_extra_context_values"] = extra_context_values
            rows.append(record)
        return rows, score, header_idx

    # Fallback: headerless rows (positional parsing)
    fallback_rows = []
    for excel_row_num, row in enumerate(values, start=1):
        parsed = _parse_row_without_headers(row)
        if not parsed:
            continue
        parsed["_source_excel_row"] = excel_row_num
        parsed["_source_sheet"] = sheet_name
        parsed["_raw_row_values"] = [str(c).strip() for c in row if c is not None and str(c).strip()]
        parsed["_extra_context_values"] = []
        fallback_rows.append(parsed)

    # Low score so any real-header sheet still wins.
    return fallback_rows, 1 if fallback_rows else 0, None


def _find_header_row(values):
    best_idx = None
    best_score = 0
    max_scan = min(30, len(values))

    for i in range(max_scan):
        row = values[i]
        tokens = {
            _norm_key(cell)
            for cell in row
            if cell is not None and str(cell).strip() != ""
        }
        if not tokens:
            continue

        score = len(tokens.intersection(ALIAS_KEY_SET))
        if any(tok in tokens for tok in ["name", "player_name", "player"]):
            score += 2
        if any(tok in tokens for tok in ["number", "no", "jersey_number"]):
            score += 2
        if any(tok in tokens for tok in ["tshirt_size", "shirt_size", "t_shirt_size", "jersey_size_no_jersey", "jersey_size"]):
            score += 1
        if any(tok in tokens for tok in ["trouser_size", "pant_size", "pant_size_no_pants"]):
            score += 1
        if any(tok in tokens for tok in ["preferred_name_on_jersey", "preferred_jersey_no"]):
            score += 2

        if score > best_score:
            best_score = score
            best_idx = i

    if best_idx is None or best_score < 3:
        return None, 0
    return best_idx, best_score


def _norm_key(value):
    raw = "".join(c.lower() if c.isalnum() else "_" for c in str(value))
    return re.sub(r"_+", "_", raw).strip("_")


def _map_raw_row(raw_row):
    normalized_lookup = {
        _norm_key(k): v
        for k, v in raw_row.items()
        if not str(k).startswith("_")
    }
    mapped = {}

    for target, aliases in HEADER_ALIASES.items():
        mapped[target] = ""
        mapped_value = _lookup_value_by_aliases(normalized_lookup, aliases)
        if mapped_value:
            mapped[target] = mapped_value

    # Hard-priority for explicit row-level gender/category columns when present.
    explicit_gender = str(raw_row.get("gender", "")).strip()
    explicit_category = str(raw_row.get("category", "")).strip()
    if explicit_gender:
        mapped["gender"] = explicit_gender
        if not mapped.get("category"):
            mapped["category"] = explicit_gender
    elif explicit_category:
        mapped["category"] = explicit_category

    # Prefer jersey/display name when provided.
    jersey_name = mapped.get("jersey_name", "").strip()
    if jersey_name:
        mapped["player_name"] = jersey_name

    # Fallback name construction from first + last names.
    if not mapped.get("player_name"):
        first = mapped.get("first_name", "").strip()
        last = mapped.get("last_name", "").strip()
        full = f"{first} {last}".strip()
        if full:
            mapped["player_name"] = full

    category_hint = _detect_category_hint(mapped, raw_row)
    mapped["tshirt_size"] = _apply_category_size_hint(mapped.get("tshirt_size", ""), category_hint)
    mapped["trouser_size"] = _apply_category_size_hint(mapped.get("trouser_size", ""), category_hint)

    return mapped


def _detect_category_hint(mapped, raw_row):
    # Explicit row-level category/gender must win over side-context text.
    explicit_values = []
    for key in ("category", "gender"):
        value = str(mapped.get(key, "")).strip()
        if value:
            explicit_values.append(value)
    explicit_hint = _category_from_text(" ".join(explicit_values))
    if explicit_hint:
        return explicit_hint

    # Fallback only when explicit category/gender is missing.
    context_values = raw_row.get("_extra_context_values", []) or []
    context_text = " ".join(str(v or "") for v in context_values)
    return _category_from_text(context_text)


def _category_from_text(text):
    raw = str(text or "").upper()
    # Normalize punctuation/variants: WOMEN'S, WOMENS, etc.
    raw = raw.replace("â€™", "").replace("'", "")
    raw = re.sub(r"[^A-Z0-9 ]+", " ", raw)
    raw = re.sub(r"\s+", " ", raw).strip()
    if not raw:
        return ""
    if "WOMEN" in raw or "WOMENS" in raw or "FEMALE" in raw:
        return "WOMENS"
    if "YOUTH" in raw or "KID" in raw or "KIDS" in raw:
        return "YOUTH"
    if "MEN" in raw or "MENS" in raw or "MALE" in raw:
        return "MENS"
    return ""

def _apply_category_size_hint(size_value, category_hint):
    size = str(size_value or "").strip().upper()
    if not size:
        return size
    if category_hint == "WOMENS":
        if size in {"WXS", "WS", "WM", "WL", "WXL", "W2XL", "W3XL", "W4XL"}:
            return size
        womens_map = {
            "XS": "WXS",
            "S": "WS",
            "M": "WM",
            "L": "WL",
            "XL": "WXL",
            "2XL": "W2XL",
            "3XL": "W3XL",
            "4XL": "W4XL",
            "SMALL": "WS",
            "MEDIUM": "WM",
            "LARGE": "WL",
            "EXTRA LARGE": "WXL",
            "XXL": "W2XL",
            "XXXL": "W3XL",
            "XXXXL": "W4XL",
        }
        return womens_map.get(size, size)
    return size


def _lookup_value_by_aliases(normalized_lookup, aliases):
    # 1) Exact key match
    for alias in aliases:
        key = _norm_key(alias)
        if key in normalized_lookup and str(normalized_lookup[key]).strip() != "":
            return str(normalized_lookup[key]).strip()

    # 2) Prefix/contains fallback for verbose headers (e.g. preferred_jersey_no_see_note_1)
    for alias in aliases:
        key = _norm_key(alias)
        for lookup_key, lookup_val in normalized_lookup.items():
            if not str(lookup_val).strip():
                continue
            if (
                lookup_key.startswith(f"{key}_")
                or f"_{key}_" in lookup_key
            ):
                return str(lookup_val).strip()

    return ""


def _is_roster_candidate(mapped_row):
    player_name = str(mapped_row.get("player_name", "")).strip()
    number = str(mapped_row.get("number", "")).strip()
    sleeve = str(mapped_row.get("sleeve_type", "")).strip()
    tshirt_size = str(mapped_row.get("tshirt_size", "")).strip()
    trouser_size = str(mapped_row.get("trouser_size", "")).strip()
    tshirt_qty = str(mapped_row.get("tshirt_qty", "")).strip()
    trouser_qty = str(mapped_row.get("trouser_qty", "")).strip()

    signals = [player_name, number, sleeve, tshirt_size, trouser_size, tshirt_qty, trouser_qty]
    if not any(signals):
        return False
    # If size/sleeve exists, treat as valid roster candidate even when name/number are blank.
    if sleeve or tshirt_size or trouser_size:
        return True
    return bool(player_name or number)


def _parse_row_without_headers(row):
    cells = ["" if cell is None else str(cell).strip() for cell in row]
    non_empty = [c for c in cells if c]
    if len(non_empty) < 4:
        return None

    # Skip obvious note/title lines.
    if len(non_empty) == 1 and len(non_empty[0]) > 20:
        return None

    # IRA-style layout without usable headers:
    # first,last,preferred,number,gender,jersey_size,qty,sleeve,pants_size,pants_qty,cap_qty
    if len(cells) >= 10:
        first = cells[0]
        last = cells[1]
        preferred = cells[2]
        number = cells[3]
        tshirt_size = cells[5]
        tshirt_qty = cells[6]
        sleeve = cells[7]
        trouser_size = cells[8]
        trouser_qty = cells[9]
        cap_qty = cells[10] if len(cells) > 10 else ""
        player_name = preferred or f"{first} {last}".strip()

        if player_name and (tshirt_size or trouser_size or sleeve):
            return {
                "player_name": player_name,
                "number": number,
                "sleeve_type": sleeve,
                "tshirt_size": tshirt_size,
                "tshirt_qty": tshirt_qty,
                "trouser_size": trouser_size,
                "trouser_qty": trouser_qty,
                "cap_qty": cap_qty,
            }

    # Minimal no-header shape:
    # player_name,number,sleeve_type,tshirt_size,tshirt_qty,trouser_size,trouser_qty
    if len(cells) >= 7:
        player_name = cells[0]
        if player_name and (cells[2] or cells[3] or cells[5]):
            return {
                "player_name": player_name,
                "number": cells[1],
                "sleeve_type": cells[2],
                "tshirt_size": cells[3],
                "tshirt_qty": cells[4],
                "trouser_size": cells[5],
                "trouser_qty": cells[6],
                "cap_qty": cells[7] if len(cells) > 7 else "",
            }

    return None


def _normalize_sleeve(value):
    v = _pick_first_option(value).upper()
    if v in {"HALF", "HALF SLEEVE", "SHORT", "SHORT SLEEVE"}:
        return "HALF"
    if v in {"FULL", "FULL SLEEVE", "LONG", "LONG SLEEVE"}:
        return "FULL"
    if v in {"3/4", "3/4TH", "3/4 TH", "THREE FOURTH", "THREE-FOURTH"}:
        return "3/4 TH"
    return v


def _normalize_size(value):
    v = _pick_first_option(value).replace("\u00a0", " ")
    v = v.strip().upper().replace("-", " ").replace("'", "")
    v = v.strip(" .'\"")
    v = re.sub(r"\s+", " ", v)
    if not v:
        return v
    # Compact common spacing variants: "X L" -> "XL", "2 XL" -> "2XL"
    v = re.sub(r"^(\d+)\s*XL$", r"\1XL", v)
    v = re.sub(r"^X\s*L$", "XL", v)
    v = re.sub(r"^XX\s*L$", "2XL", v)
    v = re.sub(r"^XXX\s*L$", "3XL", v)
    v = re.sub(r"^XXXX\s*L$", "4XL", v)
    aliases = {
        "EXTRA SMALL": "XS",
        "SMALL": "S",
        "MEDIUM": "M",
        "LARGE": "L",
        "EXTRA LARGE": "XL",
        "ADULT EXTRA SMALL": "XS",
        "ADULT SMALL": "S",
        "ADULT MEDIUM": "M",
        "ADULT LARGE": "L",
        "ADULT EXTRA LARGE": "XL",
        "MENS EXTRA SMALL": "XS",
        "MENS SMALL": "S",
        "MENS MEDIUM": "M",
        "MENS LARGE": "L",
        "MENS EXTRA LARGE": "XL",
        "MEN EXTRA SMALL": "XS",
        "MEN SMALL": "S",
        "MEN MEDIUM": "M",
        "MEN LARGE": "L",
        "MEN EXTRA LARGE": "XL",
        "XX LARGE": "2XL",
        "XXX LARGE": "3XL",
        "XXXX LARGE": "4XL",
        "ADULT XS": "XS",
        "ADULT S": "S",
        "ADULT M": "M",
        "ADULT L": "L",
        "ADULT XL": "XL",
        "ADULT 2XL": "2XL",
        "ADULT 3XL": "3XL",
        "ADULT 4XL": "4XL",
        "YOUTH XXS": "YXXS",
        "YOUTH XS": "YXS",
        "YOUTH S": "YS",
        "YOUTH M": "YM",
        "YOUTH L": "YL",
        "YOUTH XL": "YXL",
        "WOMEN XS": "WXS",
        "WOMEN S": "WS",
        "WOMEN M": "WM",
        "WOMEN L": "WL",
        "WOMEN XL": "WXL",
        "WOMEN SMALL": "WS",
        "WOMEN MEDIUM": "WM",
        "WOMEN LARGE": "WL",
        "WOMEN EXTRA LARGE": "WXL",
        "WOMENS XS": "WXS",
        "WOMENS S": "WS",
        "WOMENS M": "WM",
        "WOMENS L": "WL",
        "WOMENS XL": "WXL",
        "WOMENS SMALL": "WS",
        "WOMENS MEDIUM": "WM",
        "WOMENS LARGE": "WL",
        "WOMENS EXTRA LARGE": "WXL",
        "WOMEN 2XL": "W2XL",
        "WOMEN 3XL": "W3XL",
        "WOMEN 4XL": "W4XL",
        "WOMENS 2XL": "W2XL",
        "WOMENS 3XL": "W3XL",
        "WOMENS 4XL": "W4XL",
        "XXL": "2XL",
        "XXXL": "3XL",
        "XXXXL": "4XL",
        "YOUTH SMALL": "YS",
        "YOUTH MEDIUM": "YM",
        "YOUTH LARGE": "YL",
        "YOUTH EXTRA LARGE": "YXL",
        "6Y": "YXXS",
        "6 Y": "YXXS",
        "8Y": "YXS",
        "8 Y": "YXS",
        "10Y": "YS",
        "10 Y": "YS",
        "12Y": "YM",
        "12 Y": "YM",
        "14Y": "YL",
        "14 Y": "YL",
        "16Y": "YXL",
        "16 Y": "YXL",
        "6YR": "YXXS",
        "8YR": "YXS",
        "10YR": "YS",
        "12YR": "YM",
        "14YR": "YL",
        "16YR": "YXL",
        "6 YEARS": "YXXS",
        "8 YEARS": "YXS",
        "10 YEARS": "YS",
        "12 YEARS": "YM",
        "14 YEARS": "YL",
        "16 YEARS": "YXL",
    }
    return aliases.get(v, v)


def _parse_int_value(value):
    if value is None:
        return None
    raw = _pick_first_option(value)
    if not raw:
        return None
    raw = raw.replace(",", "")
    try:
        return int(raw)
    except ValueError:
        pass
    try:
        f = float(raw)
    except ValueError:
        return None
    if f.is_integer():
        return int(f)
    return None


def _normalize_number(value):
    raw = _pick_first_option(value)
    if not raw:
        return ""
    int_like = _parse_int_value(raw)
    if int_like is not None:
        return str(int_like)
    return raw


def _pick_first_option(value):
    raw = "" if value is None else str(value).strip()
    if not raw:
        return ""
    # For combined entries like "SHORT/LONG", "S/M", "10 or 11", pick first token.
    first = re.split(r"\s*(?:/|\\|\||,|;|\bor\b)\s*", raw, maxsplit=1)[0]
    return first.strip()


def _build_headers(values, header_idx, group_idx=None):
    header_row = values[header_idx]
    group_row = values[group_idx] if group_idx is not None and group_idx >= 0 else []
    raw_headers = []
    seen = {}

    for i, cell in enumerate(header_row):
        sub = "" if cell is None else str(cell).strip()
        grp = ""
        if i < len(group_row) and group_row[i] is not None:
            grp = str(group_row[i]).strip()

        if grp and sub:
            composed = f"{grp} {sub}".strip()
        else:
            composed = sub or grp

        if not composed:
            raw_headers.append("")
            continue

        key = _norm_key(composed)
        if key in seen:
            seen[key] += 1
            key = f"{key}_{seen[key]}"
        else:
            seen[key] = 0
        raw_headers.append(key)

    return raw_headers


def _should_shift_to_subheader(values, header_idx):
    if header_idx + 1 >= len(values):
        return False

    current_tokens = _row_tokens(values[header_idx])
    next_tokens = _row_tokens(values[header_idx + 1])
    if not next_tokens:
        return False

    current_generic = len(current_tokens.intersection({"player", "jersey", "pants", "cap", "qty"}))
    next_specific = len(
        next_tokens.intersection(
            {
                "first_name",
                "last_name",
                "preferred_name_on_jersey",
                "preferred_jersey_no",
                "pant_size_no_pants",
                "qty",
                "sleeve",
                "sleeve_type",
                "half_sleeve_full_sleeve",
            }
        )
    )
    return current_generic >= 2 and next_specific >= 2


def _row_tokens(row):
    return {
        _norm_key(cell)
        for cell in row
        if cell is not None and str(cell).strip() != ""
    }


def call_openai_for_roster(extracted):
    from openai import OpenAI

    api_key = os.environ.get("OPENAI_API_KEY")
    if not api_key:
        raise RuntimeError("OPENAI_API_KEY is not configured")

    model = os.environ.get("OPENAI_ROSTER_MODEL", "gpt-4.1")
    client = OpenAI(api_key=api_key)

    system_prompt = (
        "You are a strict roster extraction engine for sports uniform orders.\n"
        "Return JSON object with key 'players' only.\n"
        "Never invent players. Never guess missing values. Never fabricate quantities.\n"
        "Include only actual player rows from tabular data.\n"
        "Ignore title rows, notes, legends, and non-roster sections.\n"
        "If value is missing, keep it blank (or null for qty fields).\n\n"
        "Output row fields:\n"
        "- row_number\n"
        "- player_name\n"
        "- print_name\n"
        "- number\n"
        "- sleeve_type\n"
        "- tshirt_size\n"
        "- tshirt_qty\n"
        "- trouser_size\n"
        "- trouser_qty\n\n"
        "Normalization rules:\n"
        "- sleeve aliases: SHORT/HALF -> HALF, LONG/FULL -> FULL.\n"
        "- if sleeve contains multiple values like 'Short/Long', pick first token.\n"
        "- size aliases: SMALL->S, MEDIUM->M, LARGE->L.\n"
        "- normalize XL variants: '2XL'/'XXL' -> 2XL, '3XL'/'XXXL' -> 3XL, '4XL'/'XXXXL' -> 4XL.\n"
        "- preserve jersey number as text; keep leading zeroes if present.\n"
        "- if qty missing, keep null (downstream deterministic logic may default).\n"
        "- if tshirt size missing but trouser size exists, keep row; same vice-versa.\n"
        "- if both tshirt_size and trouser_size missing, skip row.\n"
        "- if player name is missing, keep empty string (do not invent a placeholder name).\n\n"
        "Women detection hints:\n"
        "- detect womens markers from values such as WOMEN'S, WOMENS, FEMALE, GIRLS even in side cells.\n"
        "- if row is women and size is adult-style (S/M/L/XL/2XL/3XL/4XL), convert to WS/WM/WL/WXL/W2XL/W3XL/W4XL.\n"
    )

    clipped = extracted["text"][:15000]
    user_content = (
        "Extract roster rows from this spreadsheet-like content.\n"
        "Return valid JSON only as: {\"players\": [...]}.\n"
        "No prose.\n\n"
        f"{clipped}"
    )

    response = client.chat.completions.create(
        model=model,
        messages=[
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_content},
        ],
        temperature=0,
        response_format={"type": "json_object"},
    )

    content = response.choices[0].message.content or "{}"
    payload = json.loads(content)
    players = payload.get("players", [])
    if not isinstance(players, list):
        return []
    return players

