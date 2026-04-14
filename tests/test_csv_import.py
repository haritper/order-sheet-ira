import io
from types import SimpleNamespace

from openpyxl import Workbook

from app.players.services import parse_players_ai, parse_players_csv


def test_parse_players_csv_success():
    csv_data = io.BytesIO(
        b"player_name,number,sleeve_type,tshirt_size,tshirt_qty,trouser_size,trouser_qty\n"
        b"John,10,HALF,M,2,M,2\n"
    )
    rows, errors = parse_players_csv(csv_data)
    assert len(rows) == 1
    assert not errors


def test_parse_players_csv_validation_error():
    csv_data = io.BytesIO(
        b"player_name,number,sleeve_type,tshirt_size,tshirt_qty,trouser_size,trouser_qty\n"
        b",10,BAD,M,0,M,1\n"
    )
    rows, errors = parse_players_csv(csv_data)
    assert not rows
    assert errors


def test_parse_players_ai_csv_deterministic_no_defaults():
    file_obj = SimpleNamespace(
        filename="roster.csv",
        read=lambda: (
            b"name,no,sleeve type,t shirt size,t shirt qty,trouser size,trouser qty\n"
            b"Advaya,12,HALF,M,1,M,1\n"
        ),
    )
    rows, errors = parse_players_ai(file_obj)
    assert not errors
    assert len(rows) == 1
    assert rows[0]["player_name"] == "Advaya"
    assert rows[0]["number"] == "12"


def test_parse_players_ai_xlsx_deterministic_alias_headers():
    wb = Workbook()
    ws = wb.active
    ws.append(["Player", "No", "Sleeve Type", "T Shirt Size", "T Shirt Qty", "Trouser Size", "Trouser Qty"])
    ws.append(["Krish", "77", "FULL", "L", 2, "L", 2])
    stream = io.BytesIO()
    wb.save(stream)
    payload = stream.getvalue()

    file_obj = SimpleNamespace(filename="roster.xlsx", read=lambda: payload)
    rows, errors = parse_players_ai(file_obj)
    assert not errors
    assert len(rows) == 1
    assert rows[0]["player_name"] == "Krish"
    assert rows[0]["number"] == "77"


def test_parse_players_ai_xlsx_detects_header_not_first_row():
    wb = Workbook()
    ws = wb.active
    ws.title = "Order Summary"
    ws.append(["IRA SPORTSWEAR ORDER"])
    ws.append(["Generated On", "2026-03-17"])
    ws.append([])
    ws.append(["Player Name", "No", "Sleeve Type", "T Shirt Size", "T Shirt Qty", "Trouser Size", "Trouser Qty"])
    ws.append(["Advaya", "10", "HALF", "M", 1, "M", 1])
    ws.append(["Avi", "15", "FULL", "L", 2, "L", 2])

    stream = io.BytesIO()
    wb.save(stream)
    payload = stream.getvalue()

    file_obj = SimpleNamespace(filename="roster.xlsx", read=lambda: payload)
    rows, errors = parse_players_ai(file_obj)
    assert not errors
    assert len(rows) == 2
    assert rows[0]["player_name"] == "Advaya"
    assert rows[1]["player_name"] == "Avi"


def test_parse_players_ai_xlsx_picks_best_sheet():
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Summary"
    ws1.append(["Notes only"])
    ws1.append(["No roster here"])

    ws2 = wb.create_sheet("Roster")
    ws2.append(["Player", "No", "Sleeve", "Shirt Size", "Shirt Qty", "Pant Size", "Pant Qty"])
    ws2.append(["Krish", "7", "HALF", "M", 1, "M", 1])

    stream = io.BytesIO()
    wb.save(stream)
    payload = stream.getvalue()

    file_obj = SimpleNamespace(filename="multi.xlsx", read=lambda: payload)
    rows, errors = parse_players_ai(file_obj)
    assert not errors
    assert len(rows) == 1
    assert rows[0]["player_name"] == "Krish"


def test_parse_players_ai_xlsx_ira_style_grouped_headers():
    wb = Workbook()
    ws = wb.active
    ws.title = "Uniform Order Count"

    ws.append(["SDCCYA Uniform Order U13 for ATL"])
    ws.append(["Please Refer Size Chart"])
    ws.append([])
    ws.append(
        [
            "Player",
            "Player",
            "Preferred Name on Jersey",
            "Preferred Jersey No",
            "Gender",
            "Jersey Size / No Jersey",
            "Quantity",
            "Half Sleeve / Full Sleeve",
            "Pants",
            "Pants",
            "Cap",
        ]
    )
    ws.append(
        [
            "First Name",
            "Last Name",
            "",
            "",
            "",
            "",
            "",
            "",
            "Pant Size/ No Pants",
            "Qty",
            "Qty",
        ]
    )
    ws.append(["Advaya", "Bhumkar", "Advaya", "54", "Male", "Adult S", 2, "Full Sleeve", "Youth XL", 1, 1])
    ws.append(["Avi", "Bagrodia", "Avi", "183", "Male", "Adult XS", 2, "Half Sleeve", "Youth XL", 1, 0])

    stream = io.BytesIO()
    wb.save(stream)
    payload = stream.getvalue()

    file_obj = SimpleNamespace(filename="ira_style.xlsx", read=lambda: payload)
    rows, errors = parse_players_ai(file_obj)

    assert not errors
    assert len(rows) == 2
    assert rows[0]["player_name"] == "Advaya"
    assert rows[0]["number"] == "54"
    assert rows[0]["sleeve_type"] == "FULL"
    assert rows[0]["tshirt_size"] == "S"
    assert rows[0]["trouser_size"] == "YXL"


def test_parse_players_ai_rejects_unsupported_type():
    file_obj = SimpleNamespace(filename="roster.pdf", read=lambda: b"%PDF")
    try:
        parse_players_ai(file_obj)
        assert False, "Expected ValueError for unsupported file type"
    except ValueError as exc:
        assert ".xlsx or .csv" in str(exc)


def test_parse_players_ai_xlsx_ira_style_note_number_and_float_qty():
    wb = Workbook()
    ws = wb.active
    ws.title = "Uniform Order Count"

    ws.append(["SDCCYA Uniform Order U13 for ATL"])
    ws.append(["Please Refer Size Chart"])
    ws.append([])
    ws.append(
        [
            "Player",
            "Player",
            "Jersey",
            "Preferred Jersey No*See Note 1 above*",
            "Gender",
            "Jersey Size / No Jersey",
            "Quantity",
            "Half Sleeve / Full Sleeve",
            "Pants",
            "Pants",
            "Cap",
        ]
    )
    ws.append(["First Name", "Last Name", "Preferred Name on Jersey", "", "", "", "", "", "Pant Size/ No Pants", "Qty", "Qty"])
    ws.append(["Advaya", "Bhumkar", "Advaya", "54", "Male", "Adult S", 2.0, "Full Sleeve", "Youth XL", 1.0, 1.0])

    stream = io.BytesIO()
    wb.save(stream)
    payload = stream.getvalue()

    file_obj = SimpleNamespace(filename="ira_note_float.xlsx", read=lambda: payload)
    rows, errors = parse_players_ai(file_obj)

    assert not errors
    assert len(rows) == 1
    assert rows[0]["number"] == "54"
    assert rows[0]["tshirt_qty"] == 2
    assert rows[0]["trouser_qty"] == 1


def test_parse_players_ai_xlsx_without_headers_positional_fallback():
    wb = Workbook()
    ws = wb.active
    ws.title = "No Headers"

    ws.append(["SDCCYA Uniform Order U13 for ATL"])
    ws.append(["Please refer size chart"])
    ws.append([])
    ws.append(["Advaya", "Bhumkar", "Advaya", "54", "Male", "Adult S", 2, "Full Sleeve", "Youth XL", 1, 1])
    ws.append(["Avi", "Bagrodia", "Avi", "183", "Male", "Adult XS", 2, "Half Sleeve", "Youth XL", 1, 0])

    stream = io.BytesIO()
    wb.save(stream)
    payload = stream.getvalue()

    file_obj = SimpleNamespace(filename="no_headers.xlsx", read=lambda: payload)
    rows, errors = parse_players_ai(file_obj)

    assert not errors
    assert len(rows) == 2
    assert rows[0]["player_name"] == "Advaya"
    assert rows[0]["number"] == "54"
    assert rows[0]["sleeve_type"] == "FULL"
    assert rows[0]["tshirt_size"] == "S"
    assert rows[0]["trouser_size"] == "YXL"


def test_parse_players_ai_csv_without_headers_positional_fallback():
    file_obj = SimpleNamespace(
        filename="roster.csv",
        read=lambda: (
            b"Advaya,54,HALF,M,1,M,1\n"
            b"Avi,183,FULL,L,2,L,2\n"
        ),
    )
    rows, errors = parse_players_ai(file_obj)
    assert not errors
    assert len(rows) == 2
    assert rows[0]["player_name"] == "Advaya"
    assert rows[0]["number"] == "54"


def test_parse_players_ai_tshirt_only_defaults_missing_sleeve_to_half():
    file_obj = SimpleNamespace(
        filename="roster.csv",
        read=lambda: (
            b"player_name,number,sleeve_type,tshirt_size,tshirt_qty,trouser_size,trouser_qty\n"
            b"Only Tee,11,,M,2,,\n"
        ),
    )
    rows, errors = parse_players_ai(file_obj)
    assert not errors
    assert len(rows) == 1
    assert rows[0]["player_name"] == "Only Tee"
    assert rows[0]["sleeve_type"] == "HALF"
    assert rows[0]["tshirt_size"] == "M"
    assert rows[0]["tshirt_qty"] == 2
    assert rows[0]["trouser_size"] == ""
    assert rows[0]["trouser_qty"] == 0


def test_parse_players_ai_trouser_only_row_is_allowed():
    file_obj = SimpleNamespace(
        filename="roster.csv",
        read=lambda: (
            b"player_name,number,sleeve_type,tshirt_size,tshirt_qty,trouser_size,trouser_qty\n"
            b"Only Trouser,21,,,,L,1\n"
        ),
    )
    rows, errors = parse_players_ai(file_obj)
    assert not errors
    assert len(rows) == 1
    assert rows[0]["player_name"] == "Only Trouser"
    assert rows[0]["tshirt_size"] == ""
    assert rows[0]["tshirt_qty"] == 0
    assert rows[0]["trouser_size"] == "L"
    assert rows[0]["trouser_qty"] == 1


def test_parse_players_ai_row_with_both_sizes_missing_is_skipped():
    file_obj = SimpleNamespace(
        filename="roster.csv",
        read=lambda: (
            b"player_name,number,sleeve_type,tshirt_size,tshirt_qty,trouser_size,trouser_qty\n"
            b"Skip Me,31,HALF,,,,\n"
        ),
    )
    rows, errors = parse_players_ai(file_obj)
    assert not errors
    assert len(rows) == 0


def test_parse_players_ai_womens_side_marker_in_unlabeled_column():
    wb = Workbook()
    ws = wb.active
    ws.title = "North Forsyth Cricket Club"

    ws.append(["North Forsyth Cricket Club"])
    ws.append([])
    ws.append(["Name", "Jersey Size", "Long or Short Sleeve", "Pant Size", "Jersey Number", "Name on Jersey", ""])
    ws.append(["NFCC", "M", "HALF", "M", "6", "Sanvi", "WOMEN'S"])

    stream = io.BytesIO()
    wb.save(stream)
    payload = stream.getvalue()

    file_obj = SimpleNamespace(filename="north_forsyth.xlsx", read=lambda: payload)
    rows, errors = parse_players_ai(file_obj)

    assert not errors
    assert len(rows) == 1
    assert rows[0]["tshirt_size"] == "WM"
    assert rows[0]["trouser_size"] == "WM"


def test_parse_players_ai_womens_side_marker_with_curly_apostrophe():
    wb = Workbook()
    ws = wb.active
    ws.title = "North Forsyth Cricket Club"

    ws.append(["North Forsyth Cricket Club"])
    ws.append([])
    ws.append(["Name", "Jersey Size", "Long or Short Sleeve", "Pant Size", "Jersey Number", "Name on Jersey", ""])
    ws.append(["NFCC", "M", "HALF", "M", "6", "Sanvi", "WOMEN’S"])

    stream = io.BytesIO()
    wb.save(stream)
    payload = stream.getvalue()

    file_obj = SimpleNamespace(filename="north_forsyth_curly.xlsx", read=lambda: payload)
    rows, errors = parse_players_ai(file_obj)

    assert not errors
    assert len(rows) == 1
    assert rows[0]["tshirt_size"] == "WM"
    assert rows[0]["trouser_size"] == "WM"


def test_parse_players_ai_womens_marker_maps_word_sizes_to_w_prefix():
    wb = Workbook()
    ws = wb.active
    ws.title = "Women Rows"

    ws.append(["Women Rows"])
    ws.append([])
    ws.append(["Name", "Jersey Size", "Long or Short Sleeve", "Pant Size", "Jersey Number", "Name on Jersey", ""])
    ws.append(["NFCC", "Large", "HALF", "Medium", "6", "Sanvi", "WOMEN'S"])

    stream = io.BytesIO()
    wb.save(stream)
    payload = stream.getvalue()

    file_obj = SimpleNamespace(filename="women_rows.xlsx", read=lambda: payload)
    rows, errors = parse_players_ai(file_obj)

    assert not errors
    assert len(rows) == 1
    assert rows[0]["tshirt_size"] == "WL"
    assert rows[0]["trouser_size"] == "WM"
